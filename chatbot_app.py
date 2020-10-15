import tkinter as tk
import integration_wit
import openpyxl


class ChatbotApp:
    def __init__(self, master, messages, credit_qst):
        self.master = master
        self.question = 0
        self.messages = messages
        self.credit_qst = credit_qst

        # Create chat window and first message
        self.txt_chatbox = tk.Text(self.master, bd=0, bg="white", font=("Verdana", 12))
        self.txt_chatbox.tag_configure("bold", font=("Verdana", 12, "bold"), spacing3=11)
        self.txt_chatbox.insert(tk.END, *messages["welcome"])
        self.txt_chatbox.config(state=tk.DISABLED, wrap=tk.WORD)

        # Bind scrollbar to chat window
        self.scrollbar = tk.Scrollbar(self.master, command=self.txt_chatbox.yview)
        self.txt_chatbox['yscrollcommand'] = self.scrollbar.set

        # Create Button to send message
        self.btn_send = tk.Button(self.master, font=("Verdana", 12, 'bold'), text="Wyślij", width=10, height=5,
                                  bd=0, bg="#fbdb44", activebackground="#f0d058", fg='#000000',
                                  anchor=tk.CENTER, cursor="hand2",
                                  command=self.send_wit)
        self.master.bind('<Return>', self.btn_send['command'])  # Bind enter key to a command in the button

        # Create the box to enter message
        self.ent_entrybox = tk.Text(self.master, bd=0, bg="white", font=("Verdana", 12), spacing3=11)
        self.ent_entrybox.config(wrap=tk.WORD)
        self.ent_entrybox.pack()

        # Place and size all components on the screen
        self.scrollbar.place(x=380, y=6, height=385)
        self.txt_chatbox.place(x=6, y=6, height=386, width=375)
        self.ent_entrybox.place(x=128, y=401, height=90, width=263)
        self.btn_send.place(x=6, y=401, height=90)

    def user_msg(self):
        msg = self.ent_entrybox.get("1.0", 'end-1c').strip()  # Gets text from the textbox
        self.ent_entrybox.delete("0.0", tk.END)  # Deletes users text
        self.txt_chatbox.config(state=tk.NORMAL, wrap=tk.WORD, spacing3=11)
        self.txt_chatbox.insert(tk.END, *self.messages["you"], msg + '\n')
        self.txt_chatbox.see(tk.END)
        return msg

    def chatbot_msg(self, text, additional=''):
        self.txt_chatbox.insert(tk.END, *self.messages[text], additional)
        self.txt_chatbox.config(state=tk.DISABLED, wrap=tk.WORD, spacing3=11)
        self.txt_chatbox.see(tk.END)

    def send_wit(self):
        msg = self.user_msg()
        if msg == '':
            self.chatbot_msg("empty_msg")
        else:
            res = integration_wit.wit_response(msg)
            if res == 'agreement':
                self.chatbot_msg("chatbot", self.credit_qst[self.question])
                self.handle_response()
            elif res == "resistance":
                self.chatbot_msg("help")
            else:
                self.chatbot_msg(res)
                self.update_excel("types", res)

    def handle_response(self):
        self.btn_send['command'] = self.send_data

    def connect_wit(self):
        self.btn_send['command'] = self.send_wit

    def send_data(self):
        msg = self.user_msg()

        if msg == '':
            self.chatbot_msg("empty_msg")
        else:
            if self.question == 0 and msg.lower() == "tak":
                self.chatbot_msg("regular_client")
                self.connect_wit()
            elif self.question == self.credit_qst.index(self.credit_qst[-1]):
                self.check_form(msg)
            else:
                self.question += 1
                self.chatbot_msg("chatbot", self.credit_qst[self.question])

    def check_form(self, msg):
        if msg.lower() == "tak":
            self.chatbot_msg("thanks")
            self.update_excel("statuses", "success")

        else:
            self.chatbot_msg("help")
            self.update_excel("statuses", "failure")

    def update_excel(self, sheet_name, text):
        file = 'data.xlsx'
        wb = openpyxl.load_workbook(file)
        ws = wb[sheet_name]
        for row in ws.iter_rows(1):
            for cell in row:
                if cell.value == text:
                    ws.cell(row=cell.row, column=cell.column+2).value += 1
                    wb.save(file)
        self.connect_wit()

